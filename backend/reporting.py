"""
Enhanced Reporting Module - Excel export, PDF generation, and scheduled reports
"""
import os
import pandas as pd
from datetime import datetime, timedelta
from typing import Dict, List, Optional
import json

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    HAS_REPORTLAB = False  # Simplified for now
except ImportError:
    HAS_REPORTLAB = False

BACKEND_DIR = os.path.dirname(os.path.abspath(__file__))
REPORTS_DIR = os.path.join(BACKEND_DIR, "data", "reports")
os.makedirs(REPORTS_DIR, exist_ok=True)


class ReportGenerator:
    """Generate comprehensive reports in various formats"""
    
    @staticmethod
    def generate_inventory_report(df: pd.DataFrame, include_charts: bool = True) -> bytes:
        """Generate inventory report in Excel format"""
        if not HAS_OPENPYXL or df is None or df.empty:
            return None
        
        try:
            wb = Workbook()
            
            # Remove default sheet and create new ones
            if 'Sheet' in wb.sheetnames:
                del wb['Sheet']
            
            # Sheet 1: Summary
            ws_summary = wb.create_sheet('Summary', 0)
            ReportGenerator._add_summary_sheet(ws_summary, df)
            
            # Sheet 2: Inventory Details
            ws_inventory = wb.create_sheet('Inventory')
            ReportGenerator._add_inventory_sheet(ws_inventory, df)
            
            # Sheet 3: Low Stock Alert
            ws_low_stock = wb.create_sheet('Low Stock')
            ReportGenerator._add_low_stock_sheet(ws_low_stock, df)
            
            # Sheet 4: Category Analysis
            ws_category = wb.create_sheet('Category Analysis')
            ReportGenerator._add_category_sheet(ws_category, df)
            
            # Save to bytes
            from io import BytesIO
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            return output.getvalue()
        
        except Exception as e:
            print(f"Error generating Excel report: {e}")
            return None
    
    @staticmethod
    def _add_summary_sheet(ws, df: pd.DataFrame):
        """Add summary sheet"""
        ws['A1'] = "IntelliStock Inventory Summary Report"
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:F1')
        
        ws['A3'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        
        # KPIs
        row = 5
        ws[f'A{row}'] = "Key Performance Indicators"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        
        row += 2
        metrics = [
            ('Total Products', len(df)),
            ('Total Stock Value', f"₹{(df['selling_price'] * df['stock']).sum():,.0f}"),
            ('Total Monthly Revenue', f"₹{(df['selling_price'] * df['monthly_sales']).sum():,.0f}"),
            ('Average Stock Level', f"{df['stock'].mean():.0f}"),
            ('Low Stock Items', len(df[df['stock'] < 20])),
            ('Average Margin %', f"{((df['selling_price'] - df['cost_price']) / df['selling_price'] * 100).mean():.1f}%")
        ]
        
        for metric, value in metrics:
            ws[f'A{row}'] = metric
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
    
    @staticmethod
    def _add_inventory_sheet(ws, df: pd.DataFrame):
        """Add inventory details sheet"""
        ws['A1'] = "Complete Inventory"
        ws['A1'].font = Font(size=12, bold=True)
        
        # Add data
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                if r_idx == 1:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    @staticmethod
    def _add_low_stock_sheet(ws, df: pd.DataFrame):
        """Add low stock alert sheet"""
        ws['A1'] = "Low Stock Alert (< 20 units)"
        ws['A1'].font = Font(size=12, bold=True, color='FF0000')
        
        low_stock = df[df['stock'] < 20].sort_values('stock')
        
        if not low_stock.empty:
            for r_idx, row in enumerate(dataframe_to_rows(low_stock, index=False, header=True), 1):
                for c_idx, value in enumerate(row, 1):
                    cell = ws.cell(row=r_idx, column=c_idx, value=value)
                    if r_idx == 1:
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
                    elif c_idx == 2:  # Highlight stock column
                        cell.fill = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')
    
    @staticmethod
    def _add_category_sheet(ws, df: pd.DataFrame):
        """Add category analysis sheet"""
        ws['A1'] = "Category Analysis"
        ws['A1'].font = Font(size=12, bold=True)
        
        category_analysis = df.groupby('category').agg({
            'product_name': 'count',
            'stock': 'sum',
            'selling_price': 'mean',
            'monthly_sales': 'sum'
        }).rename(columns={'product_name': 'Products'})
        
        for r_idx, row in enumerate(dataframe_to_rows(category_analysis, index=True, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                if r_idx == 1:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
    
    @staticmethod
    def export_to_csv(df: pd.DataFrame, filename: str) -> bool:
        """Export dataframe to CSV"""
        try:
            filepath = os.path.join(REPORTS_DIR, filename)
            df.to_csv(filepath, index=False)
            return True
        except Exception as e:
            print(f"Error exporting to CSV: {e}")
            return False
    
    @staticmethod
    def generate_performance_report(
        df: pd.DataFrame,
        discount_history: List[Dict],
        price_history: List[Dict]
    ) -> Dict:
        """Generate performance analysis report"""
        try:
            report = {
                'generated_at': datetime.now().isoformat(),
                'summary': {
                    'total_products': len(df),
                    'total_stock_value': float((df['selling_price'] * df['stock']).sum()),
                    'total_revenue_30d': float((df['selling_price'] * df['monthly_sales']).sum()),
                    'avg_margin_pct': float(((df['selling_price'] - df['cost_price']) / df['selling_price'] * 100).mean())
                },
                'top_performers': {
                    'by_revenue': df.nlargest(5, 'monthly_sales')[['product_name', 'monthly_sales']].to_dict('records'),
                    'by_stock': df.nlargest(5, 'stock')[['product_name', 'stock']].to_dict('records'),
                    'by_margin': df.nlargest(5, 'selling_price')[['product_name', 'selling_price']].to_dict('records')
                },
                'risk_analysis': {
                    'low_stock_items': len(df[df['stock'] < 20]),
                    'zero_stock_items': len(df[df['stock'] == 0]),
                    'high_stock_items': len(df[df['stock'] > df['monthly_sales'] * 6])
                },
                'price_changes': len(price_history),
                'discount_changes': len(discount_history)
            }
            
            return report
        except Exception as e:
            print(f"Error generating performance report: {e}")
            return {}
    
    @staticmethod
    def save_report(report_name: str, report_data: Dict) -> bool:
        """Save report as JSON"""
        try:
            filepath = os.path.join(REPORTS_DIR, f"{report_name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json")
            with open(filepath, 'w') as f:
                json.dump(report_data, f, indent=2)
            return True
        except Exception:
            return False
    
    @staticmethod
    def get_recent_reports(limit: int = 10) -> List[str]:
        """Get list of recent reports"""
        try:
            files = []
            for f in os.listdir(REPORTS_DIR):
                if f.endswith('.json'):
                    files.append(f)
            
            files.sort(reverse=True)
            return files[:limit]
        except Exception:
            return []


class ScheduledReportConfig:
    """Configuration for scheduled reports"""
    
    CONFIG_FILE = os.path.join(REPORTS_DIR, 'schedule.json')
    
    @staticmethod
    def create_schedule(
        report_type: str,
        frequency: str,  # 'daily', 'weekly', 'monthly'
        recipients: List[str],
        include_pdf: bool = False
    ) -> bool:
        """Create scheduled report"""
        try:
            schedule = {
                'report_type': report_type,
                'frequency': frequency,
                'recipients': recipients,
                'include_pdf': include_pdf,
                'created_at': datetime.now().isoformat(),
                'next_run': (datetime.now() + timedelta(days=1)).isoformat(),
                'enabled': True
            }
            
            schedules = []
            if os.path.exists(ScheduledReportConfig.CONFIG_FILE):
                with open(ScheduledReportConfig.CONFIG_FILE, 'r') as f:
                    schedules = json.load(f)
            
            schedules.append(schedule)
            
            with open(ScheduledReportConfig.CONFIG_FILE, 'w') as f:
                json.dump(schedules, f, indent=2)
            
            return True
        except Exception:
            return False
    
    @staticmethod
    def get_schedules() -> List[Dict]:
        """Get all schedules"""
        try:
            if os.path.exists(ScheduledReportConfig.CONFIG_FILE):
                with open(ScheduledReportConfig.CONFIG_FILE, 'r') as f:
                    return json.load(f)
            return []
        except Exception:
            return []
